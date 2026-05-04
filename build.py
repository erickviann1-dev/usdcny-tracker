"""
Build script — fetches data, runs analytics, dumps to web/data.json.
Run this any time you want to refresh the dashboard.

Usage:  python build.py
Output: web/data.json
"""

import json
import sys
import os
import warnings
from pathlib import Path
from datetime import datetime, timezone

# Timezone-aware "as of" timestamps (Python 3.9+ stdlib)
try:
    from zoneinfo import ZoneInfo
except ImportError:           # very old Python
    ZoneInfo = None

warnings.filterwarnings("ignore")

# Force UTF-8 output on Windows (otherwise GBK breaks on ✓ etc.)
try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

# Workaround for SSL cert path issues with yfinance/curl_cffi
os.environ.setdefault("CURL_CA_BUNDLE", "")
os.environ.setdefault("REQUESTS_CA_BUNDLE", "")

import pandas as pd
import numpy as np

sys.path.insert(0, str(Path(__file__).parent))
from data_fetcher import get_master_data, update_cfets_usdcny_1y_fwd_cache
from analytics   import (
    run_full_analysis,
    latest_snapshot,
    build_decision_layer,
    backtest_verdict,
    compute_flip_lines,
)
from tools.build_notebook import build_replication_notebook
from tools.cross_check import run_cross_checks, summarize as summarize_checks
from config import W_CARRY, W_MISPR, W_FIXING, COMPOSITE_HIGH_THRESHOLD


def df_to_records(df: pd.DataFrame, cols=None) -> list[dict]:
    """Convert DataFrame to list of {date, col1, col2, ...} dicts, JSON-safe."""
    if cols:
        df = df[[c for c in cols if c in df.columns]]
    out = df.copy()
    out.index = out.index.strftime("%Y-%m-%d")
    out = out.reset_index().rename(columns={"index": "date"})
    out = out.replace({np.nan: None})
    return out.to_dict(orient="records")


def main():
    print("→ Fetching market data (may take 30-60s)...")
    update_cfets_usdcny_1y_fwd_cache()
    master_df, quality = get_master_data()

    if master_df.empty:
        print("✗ No data fetched.")
        sys.exit(1)

    print(f"✓ Loaded {len(master_df)} rows. Running analytics...")
    df = run_full_analysis(master_df)
    snap = latest_snapshot(df)

    # Timezone-aware build timestamps. The local "%Y-%m-%d %H:%M:%S"
    # `generated_at` is whatever the runner uses (UTC on GitHub Actions,
    # local time on Windows). The browser displays `generated_at_et`
    # which is always America/New_York (auto EST/EDT).
    now_utc = datetime.now(timezone.utc)
    et_label = "ET"
    if ZoneInfo is not None:
        try:
            ny  = now_utc.astimezone(ZoneInfo("America/New_York"))
            generated_at_et = ny.strftime("%Y-%m-%d %H:%M ") + ny.tzname()  # → "EDT" or "EST"
        except Exception:
            generated_at_et = now_utc.strftime("%Y-%m-%d %H:%M UTC")
    else:
        generated_at_et = now_utc.strftime("%Y-%m-%d %H:%M UTC")

    payload = {
        "generated_at":     datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "generated_at_utc": now_utc.strftime("%Y-%m-%d %H:%M:%S UTC"),
        "generated_at_et":  generated_at_et,   # ← user-facing on dashboard
        "snapshot":     snap,
        "decision":     build_decision_layer(snap, df),     # v3.4 verdict + stance
        "cross_checks": run_cross_checks(snap, df),         # v3.4 integrity audit
        "backtest":     backtest_verdict(df),               # v3.5 D.1 verdict track record
        "flip_lines":   compute_flip_lines(snap),           # v3.5 D.2 sensitivity
        "quality":      {k: round(float(v), 3) for k, v in quality.items()},
        "series":       df_to_records(df, cols=[
            "us_2y", "cn_2y", "yield_spread",
            "usdcny", "usdcnh", "pboc_fix", "onoffshore_gap",
            "dxy", "dxy_ret",
            "raw_carry", "carry_ma20", "carry_ma60", "carry_ma120",
            "carry_pct_rank", "carry_pct_rank_2y",
            # v3.1 — money-market funding rates + hedged-carry proxy
            "shibor_1y", "us_1y", "mm_spread", "mm_carry",
            "cip_dev_pct", "usdcny_fwd_1y", "forward_premium_pct",
            "hedged_carry_proxy", "hedged_carry_pct_rank", "hedged_carry_method",
            # v3.3 — offshore CNH funding layer
            "cnh_hibor_1y", "cnh_hibor_3m", "cnh_hibor_on",
            "cnh_funding_stress", "cnh_stress_pct_rank",
            "hedged_carry_offshore",
            "cip_fair_spot", "cip_deviation",
            "reg_predicted", "reg_residual", "reg_predicted_uni", "reg_residual_uni",
            "reg_beta_spread", "reg_beta_dxy", "reg_r2", "reg_residual_z",
            "mispricing_score",
            "alpha_cny_dxy", "expected_fix",
            "fixing_bias", "fixing_bias_raw",
            "bias_20d_mean", "bias_60d_mean",
            "defense_intensity", "policy_score",
            "composite_score", "composite_score_smooth",
        ]),
        "methodology": {
            "w_carry": round(W_CARRY, 2),
            "w_mispr": round(W_MISPR, 2),
            "w_fixing": round(W_FIXING, 2),
            "high_threshold": COMPOSITE_HIGH_THRESHOLD,
            "reg_window_days": 252,
            "reg_residual_z_window_days": 252,
        },
    }

    out_path = Path(__file__).parent / "docs" / "data.json"
    out_path.parent.mkdir(exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, default=str, ensure_ascii=False)

    size_kb = out_path.stat().st_size // 1024
    print(f"✓ Wrote {out_path} ({size_kb} KB)")
    print(f"  Latest date:        {snap['date']}")
    print(f"  Composite score:    {snap['composite_score']}/100")
    print(f"  USD/CNY:            {snap['usdcny']}")
    print(f"  Raw carry:          {snap['raw_carry']}%")

    # v3.4 — surface verdict + integrity check summary in build log
    decision = payload["decision"]
    cv = decision.get("carry_verdict", {})
    ps = decision.get("policy_stance", {})
    print(f"  Carry verdict:      {cv.get('headline_en','—')}")
    print(f"  Policy stance:      {ps.get('label_en','—')}")
    summary = summarize_checks(payload["cross_checks"])
    icon = {'pass': '✓', 'warn': '⚠', 'fail': '✗'}[summary['overall']]
    print(f"  Integrity checks:   {icon} {summary['counts']['pass']} pass · "
          f"{summary['counts']['warn']} warn · {summary['counts']['fail']} fail "
          f"({summary['total']} total)")

    bt = payload.get("backtest") or {}
    bst = bt.get("stats") or {}
    py = bst.get("pct_days_yes")
    py_s = f"{float(py)*100:.1f}%" if py is not None else "—"
    pctile = bst.get("current_hedged_pctile")
    streak = bst.get("current_yes_streak", "—")
    print(f"  Verdict backtest:   %YES={py_s} · proxy hedged %ile={pctile if pctile is not None else '—'} · "
          f"YES streak={streak} · n={bst.get('n_days', '—')} · "
          f"MARGINAL={bst.get('days_marginal', '—')}")

    append_build_log(snap, quality, size_kb, len(df))

    # --- Replication notebook ---
    build_replication_notebook(snap.get("date", "unknown"), snap)

    # --- Excel export ---
    write_excel(df, snap, payload["series"])


def write_excel(df, snap, series_records):
    """Write docs/usdcny_tracker.xlsx with three sheets: series, snapshot, methodology."""
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = Workbook()

    # Sheet 1: series
    ws1 = wb.active
    ws1.title = "series"
    cols = [
        "us_2y", "cn_2y", "yield_spread",
        "usdcny", "usdcnh", "pboc_fix", "onoffshore_gap",
        "dxy", "dxy_ret",
        "raw_carry", "carry_ma20", "carry_ma60", "carry_ma120",
        "carry_pct_rank", "carry_pct_rank_2y",
        "shibor_1y", "us_1y", "mm_spread", "mm_carry",
        "cip_dev_pct", "usdcny_fwd_1y", "forward_premium_pct",
        "hedged_carry_proxy", "hedged_carry_pct_rank", "hedged_carry_method",
        # v3.3 — offshore CNH funding layer
        "cnh_hibor_1y", "cnh_hibor_3m", "cnh_hibor_on",
        "cnh_funding_stress", "cnh_stress_pct_rank",
        "hedged_carry_offshore",
        "cip_fair_spot", "cip_deviation",
        "reg_predicted", "reg_residual", "reg_predicted_uni", "reg_residual_uni",
        "reg_beta_spread", "reg_beta_dxy", "reg_r2", "reg_residual_z",
        "mispricing_score",
        "alpha_cny_dxy", "expected_fix",
        "fixing_bias", "fixing_bias_raw",
        "bias_20d_mean", "bias_60d_mean",
        "defense_intensity", "policy_score",
        "composite_score", "composite_score_smooth",
    ]
    available = [c for c in cols if c in df.columns]
    export_df = df[available].copy()
    export_df.index = export_df.index.strftime("%Y-%m-%d")
    export_df.index.name = "date"
    export_df = export_df.reset_index()

    for r_idx, row in enumerate(dataframe_to_rows(export_df, index=False, header=True), 1):
        for c_idx, val in enumerate(row, 1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=val)

    # Sheet 2: snapshot
    ws2 = wb.create_sheet("snapshot")
    snap_keys = list(snap.keys())
    for c_idx, k in enumerate(snap_keys, 1):
        ws2.cell(row=1, column=c_idx, value=k)
        ws2.cell(row=2, column=c_idx, value=snap[k])

    # Sheet 3: methodology
    ws3 = wb.create_sheet("methodology")
    ws3.append(["Layer", "Name", "Formula", "Description"])
    ws3.append(["Layer 1", "Unhedged Raw Carry",
                "raw_carry = US_2Y − CN_2Y",
                "Nominal carry incentive before hedging costs"])
    ws3.append(["Layer 2", "Multivariate OLS Mispricing",
                "USD/CNY = α + β₁·Spread + β₂·DXY + ε (252d rolling)",
                "Residual isolates China-specific factors after controlling for DXY"])
    ws3.append(["Layer 3", "DXY-Adjusted Fixing Bias",
                "fixing_bias = pboc_fix − S × (1 + α × ΔDXY)",
                "Clean PBOC policy signal after stripping overnight DXY noise"])
    ws3.append(["Composite", "Composite Score",
                f"{W_CARRY:.0%}·carry_pct_rank + {W_MISPR:.0%}·mispricing_score + {W_FIXING:.0%}·policy_score",
                f"Weighted blend 0–100; «High» band begins at {COMPOSITE_HIGH_THRESHOLD}"])
    ws3.append(["Layer 1b", "Hedged return (market 1Y)",
                "100×[(1+UST1Y)×(F₁ᵧ/S)−(1+Shibor1Y)] if CFETS F₁ᵧ cached; else raw_carry−cip_dev%",
                "F₁ᵧ = CFETS USD/CNY 1Y all-in forward; cache grows daily. Fallback = 2Y CIP proxy"])
    ws3.append(["Layer 2+", "Residual z-score",
                "(reg_residual − rolling mean) / rolling σ over 252d",
                "In-sample standardisation of multivariate ε; charted in diagnostics panel"])

    xlsx_path = Path(__file__).parent / "docs" / "usdcny_tracker.xlsx"
    wb.save(xlsx_path)
    print(f"✓ Wrote {xlsx_path} ({xlsx_path.stat().st_size // 1024} KB)")


def append_build_log(snap, quality, size_kb, n_rows):
    """Append a structured entry to BUILD_LOG.md after each successful build."""
    log_path = Path(__file__).parent / "BUILD_LOG.md"
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    q_summary = " · ".join(
        f"{k} {'✅' if v > 0.8 else '⚠️' if v > 0.3 else '❌'}{v*100:.0f}%"
        for k, v in quality.items()
    )

    score = snap.get("composite_score", "—")
    usdcny = snap.get("usdcny", "—")
    carry = snap.get("raw_carry", "—")
    bias = snap.get("fixing_bias", "—")
    try:
        bias_pips = f"{float(bias)*10000:.0f} pips"
    except (ValueError, TypeError):
        bias_pips = "—"

    entry = (
        f"\n## {now}\n"
        f"- **Data date:** {snap.get('date', '—')}\n"
        f"- **Rows:** {n_rows} · **Payload:** {size_kb} KB\n"
        f"- **Composite Score:** {score}/100\n"
        f"- **USD/CNY:** {usdcny} · **Raw Carry:** {carry}% · **Fixing Bias:** {bias_pips}\n"
        f"- **Coverage:** {q_summary}\n"
        f"---\n"
    )

    header = "# Build Log\n\n> Auto-generated by `build.py` on each successful run.\n\n---\n"

    if log_path.exists():
        existing = log_path.read_text(encoding="utf-8")
        # Insert new entry after the header separator
        if "---\n" in existing:
            parts = existing.split("---\n", 1)
            content = parts[0] + "---\n" + entry + (parts[1] if len(parts) > 1 else "")
        else:
            content = existing + entry
    else:
        content = header + entry

    log_path.write_text(content, encoding="utf-8")
    print(f"✓ Appended build log → BUILD_LOG.md")


if __name__ == "__main__":
    main()
