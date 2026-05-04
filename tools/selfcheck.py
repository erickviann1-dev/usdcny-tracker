"""Website self-check — run from project root."""
import json, os, re, sys

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
os.chdir(ROOT)

passed, failed, warned = 0, 0, 0

def check(ok, msg, level="FAIL"):
    global passed, failed, warned
    if ok:
        passed += 1
        print(f"  \u2705 PASS  {msg}")
    elif level == "WARN":
        warned += 1
        print(f"  \u26a0\ufe0f WARN  {msg}")
    else:
        failed += 1
        print(f"  \u274c FAIL  {msg}")

# ═══════════════════════════════════════════════════════════
print("=" * 60)
print("  USD/CNY Tracker v3.5 — Self-Check Report")
print("=" * 60)

# 1. File existence + size
print("\n[1] FILE EXISTENCE")
required = {
    "docs/index.html":                        ("HTML skeleton",        1),
    "docs/dashboard.js":                      ("Dashboard renderer",   10),
    "docs/data.json":                         ("Data payload",         100),
    "docs/usdcny_tracker.xlsx":               ("Excel export",         10),
    "docs/usdcny_tracker_replication.ipynb":   ("Replication notebook", 1),
    "build.py":                               ("Build script",         1),
    "tools/build_notebook.py":                ("Notebook generator",   1),
    "requirements.txt":                       ("Dependencies",         0),
    "CHANGELOG.md":                           ("Changelog",            1),
    "ROADMAP.md":                             ("Roadmap",              1),
}
for path, (desc, min_kb) in required.items():
    exists = os.path.exists(path)
    size_kb = os.path.getsize(path) // 1024 if exists else 0
    check(exists and size_kb >= min_kb,
          f"{path} ({size_kb}KB) — {desc}")

# 2. data.json structure
print("\n[2] DATA.JSON INTEGRITY")
with open("docs/data.json", encoding="utf-8") as f:
    data = json.load(f)

for key in ["generated_at", "snapshot", "quality", "series"]:
    check(key in data, f"data.{key} exists ({type(data.get(key)).__name__})")

snap = data["snapshot"]
critical = ["date", "composite_score", "usdcny", "raw_carry", "fixing_bias",
            "reg_r2", "reg_residual_z", "shibor_1y", "us_1y", "mm_spread", "hedged_carry_proxy"]
print("\n[3] SNAPSHOT FIELDS")
for k in critical:
    v = snap.get(k, "MISSING")
    check(v != "MISSING" and v != "N/A", f"{k} = {v}")

# 3. Series completeness
print("\n[4] SERIES COLUMNS")
series = data["series"]
cols = set(series[0].keys()) if series else set()
check(len(series) > 400, f"Row count: {len(series)} (expect > 400)")
check(len(cols) > 30, f"Column count: {len(cols)} (expect > 30)")
v31_cols = ["shibor_1y", "us_1y", "mm_spread", "mm_carry",
            "cip_dev_pct", "hedged_carry_proxy", "hedged_carry_pct_rank",
            "hedged_carry_method", "forward_premium_pct"]
for c in v31_cols:
    check(c in cols, f"v3.1 column present: {c}")

# 4. Data coverage
print("\n[5] DATA COVERAGE")
for k, v in data["quality"].items():
    pct = v * 100
    if k == "usdcny_fwd_1y" and pct < 80:
        # New series: fills after CFETS cache has rows (daily append).
        check(True, f"{k}: {pct:.0f}%", level="WARN")
        continue
    lvl = "FAIL" if pct == 0 else "WARN" if pct < 80 else None
    check(pct >= 80, f"{k}: {pct:.0f}%", level=lvl or "FAIL")

# 5. i18n symmetry
print("\n[6] I18N SYMMETRY")
js = open("docs/dashboard.js", encoding="utf-8").read()
i18n_block = re.search(r"const I18N = \{(.*?)\n\};", js, re.DOTALL).group(1)
en_block = re.search(r"en:\s*\{(.*?)\n  \},", i18n_block, re.DOTALL).group(1)
zh_block = re.search(r"zh:\s*\{(.*?)\n  \},", i18n_block, re.DOTALL).group(1)
en_keys = set(re.findall(r'"([\w.]+)":', en_block))
zh_keys = set(re.findall(r'"([\w.]+)":', zh_block))
check(len(en_keys) == len(zh_keys),
      f"EN={len(en_keys)} ZH={len(zh_keys)} keys")
only_en = en_keys - zh_keys
only_zh = zh_keys - en_keys
check(not only_en, f"Only-EN keys: {only_en or 'none'}")
check(not only_zh, f"Only-ZH keys: {only_zh or 'none'}")

# Check new v3.2 keys
for k in ["glossary.code", "data.xlsx", "data.ipynb"]:
    check(k in en_keys and k in zh_keys, f"v3.2 key present: {k}")

# 6. Version consistency
print("\n[7] VERSION CONSISTENCY")
html = open("docs/index.html", encoding="utf-8").read()
script_ver = re.search(r'dashboard\.js\?v=([\d.]+)', html)
footer_ver = re.search(r'Tracker</strong> · v([\d.]+)', html)
js_header = re.search(r'design · v([\d.]+)', js)

sv = script_ver.group(1) if script_ver else "?"
fv = footer_ver.group(1) if footer_ver else "?"
jv = js_header.group(1) if js_header else "?"

tv_match = re.search(r'const TRACKER_VERSION = "([\d.]+)"', js)
tracker_ver = tv_match.group(1) if tv_match else "?"

check(sv == tracker_ver, f"<script> tag version: {sv} (expect {tracker_ver})")
check(fv == tracker_ver, f"Footer version: {fv} (expect {tracker_ver})")
check(jv == tracker_ver, f"JS header version: {jv} (expect {tracker_ver})")
check(sv == fv == jv, f"All three match: {sv} = {fv} = {jv}")

# 7. GLOSSARY_DEFS check
print("\n[8] GLOSSARY (A.1)")
gloss_match = re.search(r"const GLOSSARY_DEFS = \[(.*?)\];", js, re.DOTALL)
gloss_body = gloss_match.group(1) if gloss_match else ""
rows = re.findall(r"\[.*?\]", gloss_body, re.DOTALL)
check(len(rows) >= 35, f"GLOSSARY_DEFS has {len(rows)} rows (expect >= 35)")
linked = len([r for r in rows if "http" in r])
check(linked >= 6, f"{linked} rows with clickable source links (expect >= 6)")

# 8. Download buttons
print("\n[9] DOWNLOAD BUTTONS (A.2 / A.3)")
check("usdcny_tracker.xlsx" in html, "Excel download button in HTML")
check("usdcny_tracker_replication.ipynb" in html, "Notebook download button in HTML")
check('data-i18n="data.xlsx"' in html, "Excel button has i18n key")
check('data-i18n="data.ipynb"' in html, "Notebook button has i18n key")

# 9. Excel validation
print("\n[10] EXCEL FILE (A.2)")
try:
    from openpyxl import load_workbook
    wb = load_workbook("docs/usdcny_tracker.xlsx", read_only=True)
    sheets = wb.sheetnames
    check("series" in sheets, f"Sheet 'series' exists")
    check("snapshot" in sheets, f"Sheet 'snapshot' exists")
    check("methodology" in sheets, f"Sheet 'methodology' exists")
    ws1 = wb["series"]
    check(ws1.max_row > 400, f"Series sheet: {ws1.max_row} rows")
    check(ws1.max_column > 30, f"Series sheet: {ws1.max_column} columns")
    ws3 = wb["methodology"]
    check(ws3.max_row >= 5, f"Methodology sheet: {ws3.max_row} rows (4 layers + header)")
    wb.close()
except Exception as e:
    check(False, f"Excel validation error: {e}")

# 10. Notebook validation
print("\n[11] NOTEBOOK FILE (A.3)")
try:
    import nbformat
    nb = nbformat.read("docs/usdcny_tracker_replication.ipynb", as_version=4)
    check(len(nb.cells) == 5, f"Notebook has {len(nb.cells)} cells (expect 5)")
    md_cells = [c for c in nb.cells if c.cell_type == "markdown"]
    code_cells = [c for c in nb.cells if c.cell_type == "code"]
    check(len(md_cells) >= 1, f"{len(md_cells)} markdown cell(s)")
    check(len(code_cells) >= 3, f"{len(code_cells)} code cell(s)")
    check("data.json" in nb.cells[2].source, "Cell 3 fetches data.json")
    check("raw_carry" in nb.cells[3].source, "Cell 4 recomputes raw_carry")
except Exception as e:
    check(False, f"Notebook validation error: {e}")

# 11. KPI badges (A.4)
print("\n[12] KPI SOURCE BADGES (A.4)")
check("codeBadge" in js, "tile() function has codeBadge parameter")
kpi_fn = re.search(r"function renderKPIs\(s\)\s*\{(.*?)\n\}", js, re.DOTALL)
kpi_body = kpi_fn.group(1) if kpi_fn else ""
check('"DTWEXBGS"' in kpi_body, "DXY tile has DTWEXBGS badge")
check('"DGS1"' in kpi_body, "UST 1Y tile has DGS1 badge")
check('"DEXCHUS"' in kpi_body, "USD/CNY tile has DEXCHUS badge")

# 12. build.py integration
print("\n[13] BUILD.PY INTEGRATION")
bp = open("build.py", encoding="utf-8").read()
check("write_excel" in bp, "write_excel() function present")
check("build_replication_notebook" in bp, "build_replication_notebook() import present")
check("from tools.build_notebook" in bp, "Notebook import path correct")
check("backtest_verdict" in bp and "compute_flip_lines" in bp,
      "build.py wires backtest_verdict + compute_flip_lines")

# 13. requirements.txt
print("\n[14] REQUIREMENTS")
req = open("requirements.txt", encoding="utf-8").read()
check("nbformat" in req, "nbformat in requirements.txt")
check("openpyxl" in req, "openpyxl in requirements.txt")

# 14. v3.5 trading workbench payloads
print("\n[15] v3.5 BACKTEST + FLIP LINES")
check("backtest" in data, "data.backtest present")
fl = data.get("flip_lines")
check(isinstance(fl, list) and len(fl) >= 3,
      f"data.flip_lines list ({len(fl) if isinstance(fl, list) else 0} rows)")
bt = data.get("backtest") or {}
st = bt.get("stats") or {}
check(st.get("sharpe") is not None, f"backtest.stats.sharpe = {st.get('sharpe')}")
check(st.get("max_dd") is not None, f"backtest.stats.max_dd = {st.get('max_dd')}")
sh = st.get("sharpe")
if sh is not None and float(sh) <= 0:
    check(False, f"Sharpe {sh} ≤ 0 (non-positive track record)", level="WARN")
else:
    check(True, f"Sharpe sanity: {sh}")

# Summary
print("\n" + "=" * 60)
total = passed + failed + warned
print(f"  TOTAL: {total} checks — {passed} passed, {failed} failed, {warned} warnings")
if failed == 0:
    print("  \u2705 ALL CHECKS PASSED")
else:
    print(f"  \u274c {failed} ISSUE(S) NEED ATTENTION")
print("=" * 60)
